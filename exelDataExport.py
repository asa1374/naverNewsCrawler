# -*- coding:utf-8 -*-
import psycopg2
from openpyxl import load_workbook

database = psycopg2.connect(host='192.168.0.135', dbname='ICT_WEB_CRW', user='postgres', password='postgrespw',
                                port='5432')
cursor = database.cursor()
# data_only=True로 해줘야 수식이 아닌 값으로 받아온다.
load_wb = load_workbook("D:/dataSample.xlsx", data_only=True)

# 시트 이름으로 불러오기
load_ws = load_wb['텍스트 분석 분류단어']
#print(len(load_wb))
# 셀 주소로 값 출력
# print(load_ws['B2'].value)

for i in range(38):
    print('i  :::: ',i)
    print('i+  :::: ', i+1)
    # 셀 좌표로 값 출력
    text1 = load_ws.cell(i+1, 3).value #검색어 셀선택
    # text1 = load_ws.cell(i + 1, 2).value  # 컴코드 셀선택
    print(text1)
    text2 = text1.split(',')
    print(text2)
    for y in text2:
        # 검색어 쿼리
        str = """INSERT INTO public.com_text
                ("no", searchkeyword,com_code_no)
                VALUES((select COALESCE(MAX(no), 0) + 1 as rownum from com_text), %s,%s)"""
        values = (y, i+1)

        # 컴코드 쿼리
        # str = """INSERT INTO public.com_code
        #         ("no", name,type)
        #         VALUES((select COALESCE(MAX(no), 0) + 1 as rownum from com_code), %s,%s)"""
        # if i < 17:
        #     values = (y, '사업유형')
        # else:
        #     values = (y, '평가항목')

        cursor.execute(str,values)

# Commit the transaction
database.commit()

# Close the database connection
database.close()

